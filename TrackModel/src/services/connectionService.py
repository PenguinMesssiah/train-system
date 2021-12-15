# Imports
import os, sys
import pylightxl as xl
import xlsxwriter
from openpyxl import load_workbook

from PyQt5.QtCore           import pyqtSlot
from models.trackBlock      import trackBlock
from models.station         import station
from models.trackSwitch     import trackSwitch

#Service for Sending Information to Neighboring Modules

def retrieveBlockOccupancy(blockNumberList: list) -> dict:
    occupancyList = dict()
    #Opening Database
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    index = 0
    for row in database.values:
        index += 1
        if(row[0] == 'Block'):
            for curBlockNumber in blockNumberList:
                if(row[3] == curBlockNumber):
                    occupancyList.update({curBlockNumber: row[8]})

    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

    return occupancyList

def retrieveBlockList() -> list:
    blockList = []
    #Opening Database
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    index = 0
    for row in database.values:
        index += 1
        if(row[0] == 'Block'):
            curTrackBlock = trackBlock(row[1], row[2], int(row[3]), int(row[4]), int(row[5]), int(row[6]), int(row[7]), int(row[12]), row[0], int(row[8]))
            curTrackBlock.setFailureStates(int(row[9]), int(row[10]), int(row[11]))
            blockList.append(curTrackBlock)

    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

    return blockList

def setBlockOccupancyHigh(blockNumber: int):
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    for row in database.values:
        rowIndex+=1
        if(row[0] == 'Block' and row[3]==blockNumber):
            print('\nSuccessful Writing')
            database.cell(row=rowIndex, column= 8, value=1)
    
    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

def setBlockOccupancyLow(blockNumber: int):
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    for row in database.values:
        rowIndex+=1
        if(row[0] == 'Block' and row[3]==blockNumber):
            database.cell(row=rowIndex, column= 8, value=0)
    
    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

def setStationOccupancyHigh(name: str):
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    for row in database.values:
        rowIndex+=1
        if(row[0] == 'Station' and row[1]==name):
            database.cell(row=rowIndex, column= 2, value=1)
    
    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

def setStationOccupancyLow(name: str):
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    for row in database.values:
        rowIndex+=1
        if(row[0] == 'Station' and row[1]==station):
            database.cell(row=rowIndex, column= 2, value=0)

    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close