'''
Last Updated: October 11, 2021
@author: Will Scott
'''

# Imports
import os
import pylightxl as xl
import xlsxwriter
import random
from openpyxl import load_workbook

from PyQt5.QtCore           import pyqtSlot
from models.trackBlock      import trackBlock
from models.station         import station
from models.trackSwitch     import trackSwitch

#Service for Establishing and Editing the Database

# Method for Reading the Excel File containing the Track Layout
def readTrackFile() -> list:
    # Defined Variables
    trackLayout      = []
    count, rowlength = 0, 150;
    switchCount = 0;
    
    # Opening & Reading Excel File
    database = xl.readxl(fn='C:/Users/willi/Documents/School/SystemsProjectEngineering/TrackLayoutVehicleDatavF2.xlsx')
    
    for row in database.ws(ws='Green Line').rows:
        # Skipping first row to avoid the section headers
        count += 1
        if(count == 1):
            continue
        elif(count == rowlength):
            break;
        
        # Saving Block Information
        curTrackBlock = trackBlock(row[0], row[1], int(row[2]), int(row[3]), int(row[4]), int(row[5]), int(row [8]), int(row[9]), 'Block', 0)
        trackLayout.append(curTrackBlock)
        
        print('\n----------------------------Added Block----------------------')
        print("\nLine = ", curTrackBlock.line)
        print("\nSection = ", curTrackBlock.section)    
        print("\nBlockNumber = ", curTrackBlock.blockNumber)
        print("\nBlockLength = ", curTrackBlock.size)    
        print("\nBlockGrade = ", curTrackBlock.gradLevel)
        print("\nSpeedLimit = ", curTrackBlock.speedLimit)    
        print("\nElevation = ", curTrackBlock.elev)
        print("\nCum. Elevation = ", curTrackBlock.cumElev)    
        
        #Check & Adding Any Infrastructure
        if(row[6] != ''):
            infraStr = row[6].replace(' ','')
            ## Checking & Adding Switch
            if(infraStr.find('SWITCHTOYARD') != -1):
                curSwitch = trackSwitch(str(switchCount), 57, 0, 0, 0, 'Switch')
                trackLayout.append(curSwitch)
                switchCount+=1;
                
            elif(infraStr.find('SWITCHFROMYARD') != -1):
                curSwitch = trackSwitch(str(switchCount), 0, 63, 0, 0, 'Switch')
                trackLayout.append(curSwitch)
                switchCount+=1;
            
            elif(infraStr.find('SWITCH') != -1):
                infraStr = infraStr.replace('SWITCH', '')
                infraStr = infraStr[1 : len(infraStr)-1]
                
                swPosOne, swPosTwo = infraStr.rsplit(";");
                
                startBlock, endBlockOne = swPosOne.rsplit('-')
                endBlockTwo = swPosTwo.rsplit('-')[1]
                
                print("\n------------------------Added Switch---------------------------------") 
                print("\nSwitch Start Block = ", int(startBlock))
                print("\nEndBlockOne       = ", endBlockOne)
                print("\nEndBlockTwo       = ", endBlockTwo)
                
                curSwitch = trackSwitch(str(switchCount), int(startBlock), int(endBlockOne), int(endBlockTwo), 0, 'Switch')
                trackLayout.append(curSwitch)
                switchCount+=1;
                
            ## Checking & Adding Station    
            elif(infraStr.find('STATION') != -1):
                infraStr = infraStr.replace('STATION;', '')
                
                if(infraStr.find(';UNDERGROUND')):
                    infraStr = infraStr.replace(';UNDERGROUND', '')
                    curStation = station(infraStr, 0, row[1], 'Station', 1, 0)
                else:
                    curStation = station(infraStr, 0, row[1], 'Station', 0, 0)
                
                trackLayout.append(curStation)
                print("\n---------------------Added Station: ", curStation.name,"-------------------")  
        
     
    print("\nSuccessfully Finished Reading the Track Layout File")
    
    # Write to Database File & Return
    writeDatabase(trackLayout);
    return trackLayout;

# Method for Writing Failure States to the Database
def writeFailureState(blockNumber: int, failureType: int):
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    #Reading Database File
    for row in database.values:
        rowIndex+=1

        if(row[0] == 'Block' and row[3] == blockNumber):
            print('rowIndex = ', rowIndex)
            if(failureType == 0): #BR
                database.cell(row=rowIndex, column= 10, value=1)
                
            elif(failureType == 1): #PF
                database.cell(row=rowIndex, column= 11, value=1)
                
            elif(failureType == 2): #TC 
                database.cell(row=rowIndex, column= 12, value=1)       
              
    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

# Method for Writing to the Database
def writeDatabase(trackLayout: list):
    rowIndex = 0;
    
    path = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    database  = xlsxwriter.Workbook(path);
    worksheet = database.add_worksheet();

    for curObject in trackLayout:
        if(curObject.objType == 'Block'):
        
        #Writing All of the Block Info Into Single Row
            #Track Object Specific 
            worksheet.write(rowIndex, 0, curObject.objType);    
            worksheet.write(rowIndex, 1, curObject.line);    
            worksheet.write(rowIndex, 2, curObject.section)
            worksheet.write(rowIndex, 3, curObject.blockNumber);    
            worksheet.write(rowIndex, 4, curObject.size);    
            worksheet.write(rowIndex, 5, curObject.gradLevel);    
            worksheet.write(rowIndex, 6, curObject.speedLimit);    
            worksheet.write(rowIndex, 7, curObject.elev)
            worksheet.write(rowIndex, 8, curObject.occupancy)
            worksheet.write(rowIndex, 12, curObject.cumElev)
            #Failure States
            worksheet.write(rowIndex, 9, curObject.failureBR)
            worksheet.write(rowIndex, 10, curObject.failurePF)
            worksheet.write(rowIndex, 11, curObject.failureTC)

            rowIndex += 1;        
            
        elif(curObject.objType == 'Switch'):
            #Track Object Specific 
            worksheet.write(rowIndex, 0, curObject.objType);    
            worksheet.write(rowIndex, 1, curObject.elementId);  
            worksheet.write(rowIndex, 2, curObject.startBlock)
            worksheet.write(rowIndex, 3, curObject.endBlockOne);    
            worksheet.write(rowIndex, 4, curObject.endBlockTwo);    
            worksheet.write(rowIndex, 5, curObject.position);     

            rowIndex += 1;   
        
        elif(curObject.objType == 'Station'):
            #Track Object Specific 
            worksheet.write(rowIndex, 0, curObject.objType); 
            worksheet.write(rowIndex, 1, curObject.name)
            worksheet.write(rowIndex, 2, curObject.occupancy)
            worksheet.write(rowIndex, 3, curObject.section)
            worksheet.write(rowIndex, 4, curObject.underground)
            worksheet.write(rowIndex, 5, curObject.ticketSales)

            #Positioning Data
            worksheet.write(rowIndex, 12, curObject.xPos)
            worksheet.write(rowIndex, 13, curObject.yPos); 
            rowIndex += 1;   

    if(not(os.path.isfile(path))):
        database.close()


# Method for Reading the Database
def readDatabase() -> list:
    tracklayout = [];

    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    db_workbook = load_workbook(path);
    database    = db_workbook.active;
    database.title = "Track Layout Information";

    print('\n\n--------Activating Database Stream--------')
    print('\nDatabase Sheet Name = ', db_workbook.sheetnames);
    
    #Reading Database File
    for row in database.values:
        if(row[0] == 'Block'):
            curTrackBlock = trackBlock(row[1], row[2], int(row[3]), int(row[4]), int(row[5]), int(row[6]), int(row [7]), int(row[12]), row[0], int(row[8]));
            curTrackBlock.setFailureStates(row[9], row[10], row[11])
            tracklayout.append(curTrackBlock);

        elif(row[0] == 'Switch'):
            curTrackSwitch = trackSwitch(row[1], int(row[2]), int(row[3]), int(row[4]), int(row[5]), row[0]);
            tracklayout.append(curTrackSwitch);

        elif(row[0] == 'Station'):
            curStation = station(row[1], int(row[2]), row[3], 'Station', row[4], int(row[5]))
            tracklayout.append(curStation)

    print('\n--------Closing Database Stream--------')

    return tracklayout;

# Method for Adding Cartesian Grid Data for View Finder
def generatePositioningData_blueLine(trackLayout: list):
    # Variables 
    nextX = None
    nextY = None
    
    for curObject in trackLayout:
        #Loading Starting Coordinates
        if(curObject == trackLayout[0]):
            curObject.xPos = 0;
            curObject.yPos = 0;
            
            nextX = curObject.xPos + curObject.size
            nextY = curObject.yPos
            continue
        
        #Skipping Any Switches: No Positioning Need
        if(curObject.objType == 'Switch'):
            continue
        
        if(curObject.objType == 'Station'):
            curObject.xPos = nextX
            curObject.yPos = nextY
            continue
        
        if(curObject.section == 'A'):
            curObject.xPos = nextX
            curObject.yPos = nextY
             
        elif(curObject.section == 'B'):
            curObject.xPos = nextX
            curObject.yPos = nextY + 25
                
            if(curObject.objType == 'Block'):
                curObject.angle = 15
                
        elif(curObject.section == 'C'):
            if(curObject.blockNumber == 11):
                nextX = 250;
                nextY = 0;
                
            curObject.xPos = nextX
            curObject.yPos = nextY - 25
                
            if(curObject.objType == 'Block'):
                curObject.angle = -15
                
        nextX = curObject.xPos + curObject.size
        nextY = curObject.yPos
            
    writeDatabase(trackLayout); 
    
#Method for Updating Ticket Sales  
def generateTicketSales() -> int:
    #Maximum of 74 seated Passengers & Maximum of 148 standing Passengers = 222 Total
    #Arbitrarily Set 10 as the Minimum & 50 as the Maximum number of Ticket Sales per Station 
    
    path        = r"C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx"
    assert      os.path.isfile(path) 
    db_workbook = load_workbook(path);
    database    = db_workbook.active;

    rowIndex = 0
    rangeValue = random.randrange(10, 50, 3)
    #Reading Database File
    for row in database.values:
        rowIndex+=1
        random.seed(rowIndex)

        if(row[0] == 'Station'):            
            database.cell(row=rowIndex, column= 6, value=rangeValue)
            print("\nUpdated Station Occupancy to ", rangeValue )
            break;
    
    db_workbook.save(r'C:\Users\willi\eclipse-workspace\train-system\TrackModel\src\database.xlsx')
    db_workbook.close

    return rangeValue
